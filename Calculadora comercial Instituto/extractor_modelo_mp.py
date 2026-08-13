from __future__ import annotations

import json
import hashlib
import os
import re
import shutil
import sys
import zipfile
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zoneinfo import ZoneInfo

import openpyxl


BOOK = Path("Tarifas Modelo Prueba Julio 2026 - v3.xlsx")
ROOT = Path(__file__).resolve().parent
BOOK = ROOT / BOOK
CATALOG = ROOT / "tarifas_modelo_mp.js"
ERROR_REPORT = ROOT / "REPORTE_ERROR_MODELO_MP.md"
MODEL_VERSION = "2026-MP-V3"
TIME_ZONE = ZoneInfo("America/Bogota")
SHEETS = ("Tarifas Ing-Fra paquetes", "Tarifas Ing-Fra nivel a nivel")
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ERROR_TOKENS = ("#REF!", "#¡REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!")
CONDITIONS = (
    ("PRECIO_AL_PUBLICO", "Precio al público", ("D", "E", "F", "G", "H", "I", "J")),
    ("ALIANZA_MASIVA", "Alianza masiva", ("L", "M", "N", "O", "P", "Q", "R")),
    ("ALIANZA_EMPRESARIAL", "Alianza empresarial", ("T", "U", "V", "W", "X", "Y", "Z")),
    ("COLABORADOR", "Colaborador", ("AB", "AC", "AD", "AE", "AF", "AG", "AH")),
)


def qname(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    return [
        "".join(node.text or "" for node in item.iter(qname(NS_MAIN, "t")))
        for item in root.findall(qname(NS_MAIN, "si"))
    ]


def workbook_sheet_targets(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relation_map = {item.attrib["Id"]: item.attrib["Target"] for item in relationships}
    result: dict[str, str] = {}
    sheets = workbook.find(qname(NS_MAIN, "sheets"))
    if sheets is None:
        return result
    for sheet in sheets:
        relation_id = sheet.attrib[qname(NS_REL, "id")]
        target = relation_map[relation_id].replace("\\", "/")
        if target.startswith("/"):
            target = target.lstrip("/")
        elif not target.startswith("xl/"):
            target = f"xl/{target}"
        result[sheet.attrib["name"]] = target
    return result


def read_cells(archive: zipfile.ZipFile, sheet_name: str) -> dict[str, dict[str, str | None]]:
    shared = read_shared_strings(archive)
    target = workbook_sheet_targets(archive)[sheet_name]
    root = ET.fromstring(archive.read(target))
    cells: dict[str, dict[str, str | None]] = {}
    for item in root.findall(f".//{qname(NS_MAIN, 'c')}"):
        address = item.attrib["r"]
        cell_type = item.attrib.get("t")
        value_node = item.find(qname(NS_MAIN, "v"))
        formula_node = item.find(qname(NS_MAIN, "f"))
        raw = value_node.text if value_node is not None else None
        display = raw
        if cell_type == "s" and raw is not None:
            display = shared[int(raw)]
        elif cell_type == "inlineStr":
            inline = item.find(qname(NS_MAIN, "is"))
            display = (
                "".join(node.text or "" for node in inline.iter(qname(NS_MAIN, "t")))
                if inline is not None
                else ""
            )
        cells[address] = {
            "raw": raw,
            "display": display,
            "formula": formula_node.text if formula_node is not None else None,
            "type": cell_type,
        }
    return cells


def get_cell(cells: dict[str, dict[str, str | None]], address: str) -> dict[str, str | None]:
    return cells.get(address, {"raw": None, "display": None, "formula": None, "type": None})


def has_numeric_raw(cells: dict[str, dict[str, str | None]], address: str) -> bool:
    item = get_cell(cells, address)
    if item["raw"] is None or item["type"] in {"s", "inlineStr", "str", "e"}:
        return False
    try:
        Decimal(str(item["raw"]))
        return True
    except InvalidOperation:
        return False


def decimal(raw: str | None, context: str, errors: list[str]) -> Decimal | None:
    if raw is None or str(raw).strip() == "":
        errors.append(f"Campo obligatorio vacío: {context}")
        return None
    try:
        return Decimal(str(raw))
    except InvalidOperation:
        errors.append(f"Valor numérico inválido: {context}={raw!r}")
        return None


def round_cop(value: Decimal) -> int:
    return int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def normalize_discount(value: Decimal) -> str:
    normalized = value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    return format(normalized, "f").rstrip("0").rstrip(".") or "0"


def parse_plan(text: str, context: str, errors: list[str]) -> dict[str, Any] | None:
    clean = re.sub(r"\s+", " ", text).strip()
    levels = re.findall(r"\b(?:A1|A2|B1|B2|C1)\b", clean)
    hours_match = re.search(r"(\d+)\s+Horas", clean, flags=re.IGNORECASE)
    count_match = re.search(r"Plan\s+(\d+)\s+Nivel", clean, flags=re.IGNORECASE)
    if not levels or not hours_match or not count_match:
        errors.append(f"Plan no interpretable: {context}: {text!r}")
        return None
    count = int(count_match.group(1))
    hours = int(hours_match.group(1))
    if count != len(levels):
        errors.append(f"Cantidad de niveles inconsistente: {context}: declara {count}, contiene {levels}")
    return {"original": text, "levels": levels, "hours": hours, "count": count}


def inspect_package_records(cells: dict[str, dict[str, str | None]], errors: list[str]) -> list[dict[str, Any]]:
    physical: list[dict[str, Any]] = []
    current_plan: dict[str, Any] | None = None
    full_raw: str | None = None
    for row in range(5, 45):
        plan_text = get_cell(cells, f"A{row}")["display"]
        if plan_text:
            current_plan = parse_plan(str(plan_text), f"paquetes!A{row}", errors)
        raw_full = get_cell(cells, f"B{row}")["raw"]
        if raw_full is not None:
            full_raw = str(raw_full)
        if not any(has_numeric_raw(cells, f"{columns[0]}{row}") for _, _, columns in CONDITIONS):
            continue
        if current_plan is None or full_raw is None:
            errors.append(f"Plan o valor full no heredable: paquetes fila {row}")
            continue
        physical.extend(build_condition_records(cells, row, current_plan, full_raw, None, "MODELO_MP_PAQUETES", errors))
    projected: list[dict[str, Any]] = []
    for item in physical:
        projected.append({**item, "idioma": "INGLES"})
        if "C1" not in item["levels"]:
            projected.append({**item, "idioma": "FRANCES"})
    return projected


def inspect_stage_records(cells: dict[str, dict[str, str | None]], errors: list[str]) -> list[dict[str, Any]]:
    physical: list[dict[str, Any]] = []
    route_start: str | None = None
    current_plan: dict[str, Any] | None = None
    full_raw: str | None = None
    for row in range(1, 72):
        label = get_cell(cells, f"A{row}")["display"]
        if label and str(label).strip().upper().startswith("INICIO"):
            levels = re.findall(r"\b(?:A1|A2|B1|B2|C1)\b", str(label).upper())
            route_start = levels[0] if levels else None
            if route_start is None:
                errors.append(f"Ruta no interpretable: nivel a nivel!A{row}={label!r}")
            current_plan = None
            full_raw = None
        elif label and re.match(r"^\s*PLAN\s+\d+", str(label), flags=re.IGNORECASE):
            current_plan = parse_plan(str(label), f"nivel a nivel!A{row}", errors)
        raw_full = get_cell(cells, f"B{row}")["raw"]
        if raw_full is not None:
            full_raw = str(raw_full)
        if not any(has_numeric_raw(cells, f"{columns[0]}{row}") for _, _, columns in CONDITIONS):
            continue
        if current_plan is None or full_raw is None or route_start is None:
            errors.append(f"Ruta, plan o valor full no heredable: nivel a nivel fila {row}")
            continue
        physical.extend(build_condition_records(cells, row, current_plan, full_raw, route_start, "MODELO_MP_NIVEL_A_NIVEL", errors))
    projected: list[dict[str, Any]] = []
    for item in physical:
        projected.append({**item, "idioma": "INGLES"})
        if item["levels"] != ["C1"]:
            projected.append({**item, "idioma": "FRANCES"})
    return projected


def build_condition_records(
    cells: dict[str, dict[str, str | None]],
    row: int,
    plan: dict[str, Any],
    full_raw: str,
    route_start: str | None,
    model: str,
    errors: list[str],
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    full_value = decimal(full_raw, f"fila {row} valor full", errors)
    intensity_by_condition: list[tuple[str, Decimal]] = []
    for condition_id, condition_name, columns in CONDITIONS:
        payment_raw = get_cell(cells, f"{columns[0]}{row}")["raw"]
        if payment_raw is None:
            errors.append(f"Financiación incompleta: {model}, fila {row}, condición {condition_id}, número de pagos vacío")
            continue
        values: dict[str, Decimal | None] = {}
        raw_values: dict[str, str | None] = {}
        for field, column in zip(
            ("payments", "discount", "initial", "installment", "total", "hour_value", "intensity"),
            columns,
        ):
            address = f"{column}{row}"
            raw = get_cell(cells, address)["raw"]
            raw_values[field] = raw
            values[field] = decimal(raw, f"{model}!{address} ({field})", errors)
        if full_value is None or any(value is None for value in values.values()):
            continue
        payments = values["payments"]
        assert payments is not None
        if payments != payments.to_integral_value() or payments < 1:
            errors.append(f"Número de pagos inválido: {model}, fila {row}, {condition_id}: {payments}")
        item = {
            "model": model,
            "route_start": route_start,
            "row": row,
            "condition_id": condition_id,
            "condition_name": condition_name,
            "plan_original": plan["original"],
            "levels": plan["levels"],
            "hours": plan["hours"],
            "level_count": plan["count"],
            "full_raw": full_raw,
            "full": full_value,
            "raw": raw_values,
            **values,
        }
        result.append(item)
        intensity = values["intensity"]
        if intensity is not None:
            intensity_by_condition.append((condition_id, intensity))
        validate_financial_item(item, errors)
        validate_academic_item(item, errors)
    if intensity_by_condition:
        reference = intensity_by_condition[0][1]
        for condition_id, intensity in intensity_by_condition[1:]:
            if abs(intensity - reference) > Decimal("0.00000001"):
                errors.append(
                    f"Intensidad inconsistente: {model}, fila {row}, {intensity_by_condition[0][0]}={reference}, {condition_id}={intensity}"
                )
    return result


def validate_financial_item(item: dict[str, Any], errors: list[str]) -> None:
    numeric_fields = ("full", "discount", "initial", "installment", "total", "hour_value", "intensity")
    for field in numeric_fields:
        if item[field] < 0:
            errors.append(f"Valor negativo: {item['model']}, fila {item['row']}, {item['condition_id']}, {field}={item[field]}")
    if item["discount"] > 1:
        errors.append(f"Descuento superior a 100 %: {item['model']}, fila {item['row']}, {item['condition_id']}")
    authorized_zero = (
        item["model"] == "MODELO_MP_NIVEL_A_NIVEL"
        and item["route_start"] == "A1"
        and item["levels"] == ["C1"]
        and item["payments"] == 1
        and item["condition_id"] in {"ALIANZA_EMPRESARIAL", "COLABORADOR"}
        and item["discount"] == 1
        and item["initial"] == 0
        and item["installment"] == 0
        and item["total"] == 0
    )
    item["authorized_zero"] = authorized_zero
    if item["total"] == 0 and not authorized_zero:
        errors.append(f"Total $0 no autorizado: {item['model']}, fila {item['row']}, {item['condition_id']}")
    if item["discount"] == 1 and not authorized_zero:
        errors.append(f"Descuento 100 % no autorizado: {item['model']}, fila {item['row']}, {item['condition_id']}")
    if item["total"] < item["initial"]:
        errors.append(f"Total menor que cuota inicial: {item['model']}, fila {item['row']}, {item['condition_id']}")
    if item["payments"] == 1:
        if item["initial"] != item["total"]:
            errors.append(f"Contado incompleto: inicial distinta del total: {item['model']}, fila {item['row']}, {item['condition_id']}")
        if item["installment"] != 0:
            errors.append(f"Contado con cuota periódica no nula: {item['model']}, fila {item['row']}, {item['condition_id']}")
    elif item["initial"] <= 0 or item["installment"] <= 0 or item["total"] <= item["initial"]:
        errors.append(f"Financiación incompleta: {item['model']}, fila {item['row']}, {item['condition_id']}")


def validate_academic_item(item: dict[str, Any], errors: list[str]) -> None:
    tolerance = Decimal("0.00000001")
    expected_hour_value = item["total"] / Decimal(item["hours"])
    if abs(item["hour_value"] - expected_hour_value) > tolerance:
        errors.append(
            f"Horas/valor hora inconsistentes: {item['model']}, fila {item['row']}, {item['condition_id']}: "
            f"valor={item['hour_value']}, esperado={expected_hour_value} con {item['hours']} horas"
        )
    expected_intensity = Decimal(item["hours"]) / item["payments"]
    if abs(item["intensity"] - expected_intensity) > tolerance:
        errors.append(
            f"Intensidad/cálculo inconsistente: {item['model']}, fila {item['row']}, {item['condition_id']}: "
            f"valor={item['intensity']}, esperado={expected_intensity}"
        )


def make_key(item: dict[str, Any]) -> tuple[Any, ...]:
    if item["model"] == "MODELO_MP_PAQUETES":
        plan_id = "MP_INTEGRAL_" + "_".join(item["levels"])
    else:
        plan_id = f"MP_ETAPAS_INICIO_{item['route_start']}_NIVEL_{item['levels'][0]}"
    item["plan_id"] = plan_id
    return (
        item["model"],
        item["idioma"],
        plan_id,
        item["condition_id"],
        int(item["payments"]),
    )


def interface_plan_name(item: dict[str, Any]) -> str:
    levels = item["levels"]
    hours = item["hours"]
    if item["model"] == "MODELO_MP_NIVEL_A_NIVEL":
        return f"Programa por etapas – Nivel {levels[0]} ({hours} horas)"
    if levels == ["A1", "A2", "B1", "B2", "C1"]:
        return f"Plan completo de 5 niveles – A1 a C1 ({hours} horas)"
    if len(levels) == 1:
        level_text = levels[0]
    else:
        level_text = ", ".join(levels[:-1]) + f" y {levels[-1]}"
    return f"Plan de {len(levels)} niveles – {level_text} ({hours} horas)"


def panel_plan_name(item: dict[str, Any]) -> str:
    if item["model"] == "MODELO_MP_NIVEL_A_NIVEL":
        return f"Nivel {item['levels'][0]} · Ruta desde {item['route_start']} ({item['hours']} horas)"
    return interface_plan_name(item)


def normalized_record(item: dict[str, Any], updated_at: str) -> dict[str, Any]:
    plan_id = item["plan_id"]
    is_stage = item["model"] == "MODELO_MP_NIVEL_A_NIVEL"
    language_name = "Inglés" if item["idioma"] == "INGLES" else "Francés"
    raw = item["raw"]
    return {
        "modelo_tarifario_id": item["model"],
        "modelo_tarifario_version": MODEL_VERSION,
        "modalidad_cliente": "PROGRAMA_POR_ETAPAS" if is_stage else "PROGRAMA_INTEGRAL",
        "aplicacion_nacional": True,
        "idioma": item["idioma"],
        "idioma_id": item["idioma"],
        "idioma_nombre": language_name,
        "producto_id": f"{item['idioma']}_INSTITUTO",
        "zona_id": "NACIONAL",
        "zona_tarifaria": "Nacional",
        "plan_id": plan_id,
        "nombre_plan_original": item["plan_original"],
        "nombre_plan_interfaz": interface_plan_name(item),
        "nombre_plan_panel": panel_plan_name(item),
        "ruta_inicio": item["route_start"],
        "ruta_aprendizaje_interfaz": f"Ruta de aprendizaje desde {item['route_start']}" if is_stage else None,
        "nivel_contratado": item["levels"][0] if is_stage else None,
        "niveles_incluidos": list(item["levels"]),
        "numero_niveles": item["level_count"],
        "horas_academicas": item["hours"],
        "condicion_id": item["condition_id"],
        "condicion_comercial": item["condition_name"],
        "numero_pagos": int(item["payments"]),
        "porcentaje_descuento_raw_excel": str(raw["discount"]),
        "porcentaje_descuento_exacto": normalize_discount(item["discount"]),
        "valor_full_raw_excel": str(item["full_raw"]),
        "valor_full_oficial_cop": round_cop(item["full"]),
        "cuota_inicial_raw_excel": str(raw["initial"]),
        "cuota_inicial_minima_cop": round_cop(item["initial"]),
        "cuota_mensual_referencia_raw_excel": str(raw["installment"]),
        "cuota_mensual_referencia_cop": round_cop(item["installment"]),
        "valor_total_raw_excel": str(raw["total"]),
        "valor_total_oficial_cop": round_cop(item["total"]),
        "valor_por_hora_raw_excel": str(raw["hour_value"]),
        "valor_por_hora_mostrado_cop": round_cop(item["hour_value"]),
        "maxima_intensidad_mensual_raw": str(raw["intensity"]),
        "maxima_intensidad_mensual_mostrada": round_cop(item["intensity"]),
        "tipo_beneficio_continuidad": "RENOVACION_NIVEL_BONIFICADO" if item["authorized_zero"] else None,
        "archivo_origen": BOOK.name,
        "hoja_origen": SHEETS[0] if item["model"] == "MODELO_MP_PAQUETES" else SHEETS[1],
        "fila_origen": item["row"],
        "fecha_actualizacion": updated_at,
        "estado_validacion": "APROBADO",
    }


def normalized_key(item: dict[str, Any]) -> tuple[Any, ...]:
    return (
        item["modelo_tarifario_id"],
        item["idioma_id"],
        item["plan_id"],
        item["condicion_id"],
        item["numero_pagos"],
    )


def validate_normalized(records: list[dict[str, Any]]) -> None:
    problems: list[str] = []
    if len(records) != 452:
        problems.append(f"Se esperaban 452 tarifas MP y se obtuvieron {len(records)}.")
    keys = [normalized_key(item) for item in records]
    if len(keys) != len(set(keys)):
        problems.append("El catálogo normalizado contiene claves duplicadas.")
    expected = {
        ("MODELO_MP_PAQUETES", "INGLES"): 160,
        ("MODELO_MP_PAQUETES", "FRANCES"): 96,
        ("MODELO_MP_NIVEL_A_NIVEL", "INGLES"): 116,
        ("MODELO_MP_NIVEL_A_NIVEL", "FRANCES"): 80,
    }
    actual = Counter((item["modelo_tarifario_id"], item["idioma_id"]) for item in records)
    if actual != Counter(expected):
        problems.append(f"Conteos normalizados inválidos: {dict(actual)}")
    if Counter(item["idioma_id"] for item in records) != Counter({"INGLES": 276, "FRANCES": 176}):
        problems.append("Los conteos normalizados por idioma no coinciden con 276/176.")
    required = (
        "modelo_tarifario_id", "modelo_tarifario_version", "modalidad_cliente", "idioma_id", "plan_id",
        "nombre_plan_original", "nombre_plan_interfaz", "niveles_incluidos", "numero_niveles", "horas_academicas",
        "condicion_id", "condicion_comercial", "numero_pagos", "porcentaje_descuento_exacto",
        "valor_full_raw_excel", "valor_full_oficial_cop", "cuota_inicial_raw_excel", "cuota_inicial_minima_cop",
        "cuota_mensual_referencia_raw_excel", "cuota_mensual_referencia_cop", "valor_total_raw_excel",
        "valor_total_oficial_cop", "valor_por_hora_raw_excel", "valor_por_hora_mostrado_cop",
        "maxima_intensidad_mensual_raw", "maxima_intensidad_mensual_mostrada", "archivo_origen",
        "hoja_origen", "fila_origen", "fecha_actualizacion", "estado_validacion",
    )
    zero_records = []
    for index, item in enumerate(records, start=1):
        for field in required:
            if item.get(field) is None or item.get(field) == "" or item.get(field) == []:
                problems.append(f"Tarifa {index}: campo obligatorio vacío ({field}).")
        if not item["aplicacion_nacional"] or item["zona_id"] != "NACIONAL":
            problems.append(f"Tarifa {index}: la aplicación MP debe ser nacional.")
        if item["idioma_id"] == "FRANCES" and "C1" in item["niveles_incluidos"]:
            problems.append(f"Tarifa {index}: Francés no puede contener C1.")
        if item["modelo_tarifario_id"] == "MODELO_MP_NIVEL_A_NIVEL":
            if not item["ruta_inicio"] or f"INICIO_{item['ruta_inicio']}_NIVEL_{item['nivel_contratado']}" not in item["plan_id"]:
                problems.append(f"Tarifa {index}: el identificador por etapas no incluye ruta y nivel.")
        for field in (
            "valor_full_oficial_cop", "cuota_inicial_minima_cop", "cuota_mensual_referencia_cop",
            "valor_total_oficial_cop", "valor_por_hora_mostrado_cop", "maxima_intensidad_mensual_mostrada",
        ):
            if not isinstance(item[field], int) or item[field] < 0:
                problems.append(f"Tarifa {index}: entero COP/intensidad inválido ({field}).")
        if item["valor_total_oficial_cop"] == 0:
            zero_records.append(item)
    if len(zero_records) != 2 or any(item["tipo_beneficio_continuidad"] != "RENOVACION_NIVEL_BONIFICADO" for item in zero_records):
        problems.append("Los casos normalizados de $0 no coinciden con las dos continuidades C1 autorizadas.")
    institutional = ROOT / "tarifas.js"
    if not institutional.exists() or institutional.read_text(encoding="utf-8").count('"plan_id":') != 750:
        problems.append("No se confirmó que tarifas.js conserve exactamente 750 registros institucionales.")
    if problems:
        raise ValueError("\n".join(problems[:100]))


def catalog_javascript(records: list[dict[str, Any]], metadata: dict[str, Any]) -> str:
    metadata_json = json.dumps(metadata, ensure_ascii=False, indent=2)
    records_json = json.dumps(records, ensure_ascii=False, indent=2)
    return (
        "/* Archivo generado por extractor_modelo_mp.py. No editar tarifas manualmente. */\n"
        "(function (global) {\n"
        "  \"use strict\";\n"
        f"  const metadata = {metadata_json};\n"
        f"  const tarifas = {records_json};\n"
        "  global.SMART_TARIFAS_MP_META = Object.freeze(metadata);\n"
        "  global.SMART_TARIFAS_MP = Object.freeze(tarifas.map(Object.freeze));\n"
        "})(typeof window !== \"undefined\" ? window : globalThis);\n"
    )


def read_previous_records(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    text = path.read_text(encoding="utf-8")
    match = re.search(r"const tarifas = (\[.*\]);\s*global\.SMART_TARIFAS_MP_META", text, flags=re.DOTALL)
    if not match:
        return []
    return json.loads(match.group(1))


def compare_catalogs(previous: list[dict[str, Any]], current: list[dict[str, Any]]) -> dict[str, int]:
    old = {normalized_key(item): item for item in previous}
    new = {normalized_key(item): item for item in current}
    shared = old.keys() & new.keys()
    ignored = {"fecha_actualizacion"}
    modified = sum(
        1 for key in shared
        if {name: value for name, value in old[key].items() if name not in ignored}
        != {name: value for name, value in new[key].items() if name not in ignored}
    )
    return {
        "nuevas": len(new.keys() - old.keys()),
        "eliminadas": len(old.keys() - new.keys()),
        "modificadas": modified,
        "sin_cambios": len(shared) - modified,
    }


def publish_catalog(records: list[dict[str, Any]], metadata: dict[str, Any], timestamp: datetime) -> dict[str, Any]:
    previous = read_previous_records(CATALOG)
    comparison = compare_catalogs(previous, records)
    temporary = CATALOG.with_suffix(".tmp.js")
    temporary.write_text(catalog_javascript(records, metadata), encoding="utf-8", newline="\n")
    temporary_records = read_previous_records(temporary)
    validate_normalized(temporary_records)
    backup_path = None
    if CATALOG.exists():
        backup_dir = ROOT / "respaldos"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"tarifas_modelo_mp_{timestamp.strftime('%Y-%m-%d_%H%M%S')}.js"
        shutil.copy2(CATALOG, backup_path)
    os.replace(temporary, CATALOG)
    return {
        "catalogo": str(CATALOG),
        "respaldo_anterior": str(backup_path) if backup_path else None,
        "comparacion": comparison,
    }


def write_error_report(errors: list[str]) -> None:
    now = datetime.now(TIME_ZONE)
    lines = [
        "# ERROR DE EXTRACCIÓN — MODELO MP",
        "",
        f"Fecha: {now.isoformat(timespec='seconds')}",
        f"Archivo: `{BOOK.name}`",
        "",
        "El catálogo anterior no fue reemplazado y `tarifas.js` no fue modificado.",
        "",
        "## Errores",
        "",
    ]
    lines.extend(f"- {error}" for error in errors)
    ERROR_REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def audit_package_structure(archive: zipfile.ZipFile, wb_formulas: Any, wb_values: Any, errors: list[str]) -> dict[str, Any]:
    names = archive.namelist()
    external_parts = [name for name in names if "externalLink" in name]
    external_relationships: list[dict[str, str]] = []
    path_hits: list[dict[str, str]] = []
    xml_error_hits: list[dict[str, str]] = []
    for name in names:
        if not name.endswith((".xml", ".rels")):
            continue
        text = archive.read(name).decode("utf-8", "ignore")
        if name.endswith(".rels"):
            try:
                root = ET.fromstring(text)
                for relationship in root:
                    target = relationship.attrib.get("Target", "")
                    rel_type = relationship.attrib.get("Type", "")
                    target_mode = relationship.attrib.get("TargetMode", "")
                    if target_mode.lower() == "external" or "externalLink" in rel_type:
                        external_relationships.append({"part": name, "target": target, "type": rel_type})
            except ET.ParseError:
                errors.append(f"XML de relaciones inválido: {name}")
        lowered = text.lower()
        for needle in ("tarifas 2016.xlsm", "sharepoint", "onedrive", "file:///", "file://", "#ref!", "#¡ref!"):
            if needle in lowered:
                target = xml_error_hits if "ref!" in needle else path_hits
                target.append({"part": name, "needle": needle})
        if re.search(r"(?:[A-Za-z]:\\\\|[A-Za-z]:%5[cC]|\\\\\\\\[^\\])", text):
            path_hits.append({"part": name, "needle": "ruta local o UNC"})
    if external_parts:
        errors.append(f"Partes de vínculos externos: {external_parts}")
    if external_relationships:
        errors.append(f"Relaciones externas: {external_relationships}")
    if path_hits:
        errors.append(f"Rutas externas encontradas: {path_hits}")
    if xml_error_hits:
        errors.append(f"Referencias #REF! en XML: {xml_error_hits}")

    defined_names: list[dict[str, str]] = []
    for item in wb_formulas.defined_names.values():
        attr_text = str(getattr(item, "attr_text", "") or "")
        defined_names.append({"name": item.name, "value": attr_text})
        lowered = attr_text.lower()
        if any(token.lower() in lowered for token in ERROR_TOKENS) or re.search(r"\[[^\]]+\]", attr_text):
            errors.append(f"Nombre definido inválido o externo: {item.name}={attr_text}")

    formula_count = 0
    formulas_without_cache: list[str] = []
    formula_external_hits: list[str] = []
    cell_errors: list[str] = []
    for sheet_name in SHEETS:
        formula_sheet = wb_formulas[sheet_name]
        value_sheet = wb_values[sheet_name]
        for row in formula_sheet.iter_rows():
            for formula_cell in row:
                value = formula_cell.value
                cached = value_sheet[formula_cell.coordinate].value
                if formula_cell.data_type == "f":
                    formula_count += 1
                    formula = str(value)
                    if cached is None:
                        formulas_without_cache.append(f"{sheet_name}!{formula_cell.coordinate}")
                    lowered = formula.lower()
                    if re.search(r"\[[^\]]+\.(?:xlsx|xlsm|xlsb|xls)\]", formula, flags=re.IGNORECASE) or any(
                        needle in lowered for needle in ("sharepoint", "onedrive", "file://", "tarifas 2016")
                    ):
                        formula_external_hits.append(f"{sheet_name}!{formula_cell.coordinate}={formula}")
                if formula_cell.data_type == "e" or value_sheet[formula_cell.coordinate].data_type == "e":
                    cell_errors.append(f"{sheet_name}!{formula_cell.coordinate}={cached!r}")
                text = str(cached or value or "").upper()
                if any(token.upper() in text for token in ERROR_TOKENS):
                    cell_errors.append(f"{sheet_name}!{formula_cell.coordinate}={text}")
    if formulas_without_cache:
        errors.append(f"Fórmulas sin resultado almacenado ({len(formulas_without_cache)}): {formulas_without_cache[:20]}")
    if formula_external_hits:
        errors.append(f"Fórmulas con referencia externa: {formula_external_hits[:20]}")
    if cell_errors:
        errors.append(f"Errores de Excel: {cell_errors[:30]}")
    return {
        "external_parts": external_parts,
        "external_relationships": external_relationships,
        "path_hits": path_hits,
        "defined_names": defined_names,
        "formula_count": formula_count,
        "formulas_without_cache": formulas_without_cache,
        "formula_external_hits": formula_external_hits,
        "cell_errors": sorted(set(cell_errors)),
    }


def main() -> None:
    errors: list[str] = []
    wb_formulas = openpyxl.load_workbook(BOOK, data_only=False, read_only=False)
    wb_values = openpyxl.load_workbook(BOOK, data_only=True, read_only=False)
    if wb_formulas.sheetnames != list(SHEETS):
        errors.append(f"Hojas inesperadas: {wb_formulas.sheetnames}")
    with zipfile.ZipFile(BOOK) as archive:
        structure = audit_package_structure(archive, wb_formulas, wb_values, errors)
        package_cells = read_cells(archive, SHEETS[0])
        stage_cells = read_cells(archive, SHEETS[1])
    package_records = inspect_package_records(package_cells, errors)
    stage_records = inspect_stage_records(stage_cells, errors)
    records = package_records + stage_records

    keys = [make_key(item) for item in records]
    duplicates = [key for key, count in Counter(keys).items() if count > 1]
    if duplicates:
        errors.append(f"Claves duplicadas ({len(duplicates)}): {duplicates[:20]}")

    counts_model_language = Counter((item["model"], item["idioma"]) for item in records)
    expected_model_language = {
        ("MODELO_MP_PAQUETES", "INGLES"): 160,
        ("MODELO_MP_PAQUETES", "FRANCES"): 96,
        ("MODELO_MP_NIVEL_A_NIVEL", "INGLES"): 116,
        ("MODELO_MP_NIVEL_A_NIVEL", "FRANCES"): 80,
    }
    if counts_model_language != Counter(expected_model_language):
        errors.append(f"Conteos por modalidad/idioma inválidos: {dict(counts_model_language)}")
    if len(records) != 452:
        errors.append(f"Conteo MP inválido: esperado 452, obtenido {len(records)}")
    language_counts = Counter(item["idioma"] for item in records)
    if language_counts != Counter({"INGLES": 276, "FRANCES": 176}):
        errors.append(f"Conteos por idioma inválidos: {dict(language_counts)}")

    conditions = Counter(item["condition_id"] for item in records)
    zero_records = [item for item in records if item["total"] == 0]
    unauthorized_zero_records = [item for item in zero_records if not item["authorized_zero"]]
    if len(zero_records) != 2 or unauthorized_zero_records:
        errors.append(f"Casos $0 inválidos: total={len(zero_records)}, no autorizados={len(unauthorized_zero_records)}")

    target_hours = {
        address: stage_cells[address]["raw"]
        for address in ("J25", "R25", "Z25", "AH25")
    }
    if any(Decimal(str(value)) != 162 for value in target_hours.values() if value is not None):
        errors.append(f"A2 no tiene 162 horas en celdas objetivo: {target_hours}")
    target_intensity: dict[str, list[str | None]] = {}
    for row in (12, 15, 29, 32, 35, 44, 47, 50, 59, 62, 71):
        values = [stage_cells[f"{column}{row}"]["raw"] for column in ("J", "R", "Z", "AH")]
        target_intensity[str(row)] = values
        decimals = [Decimal(str(value)) for value in values if value is not None]
        if len(decimals) != 4 or any(value != decimals[0] for value in decimals[1:]):
            errors.append(f"Intensidades objetivo inconsistentes en fila {row}: {values}")

    sheet_stats = {}
    for sheet_name in SHEETS:
        sheet = wb_formulas[sheet_name]
        hidden_columns = [key for key, value in sheet.column_dimensions.items() if value.hidden]
        hidden_rows = [key for key, value in sheet.row_dimensions.items() if value.hidden]
        sheet_stats[sheet_name] = {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "merged_ranges": len(sheet.merged_cells.ranges),
            "hidden_columns": hidden_columns,
            "hidden_rows": hidden_rows,
        }

    result = {
        "status": "APROBADO" if not errors else "BLOQUEADO",
        "errors": errors,
        "sheet_stats": sheet_stats,
        "structure": structure,
        "counts": {
            "model_language": {f"{model}/{language}": count for (model, language), count in sorted(counts_model_language.items())},
            "language": dict(language_counts),
            "conditions": dict(conditions),
            "total": len(records),
            "unique_keys": len(set(keys)),
            "zero_records": len(zero_records),
        },
        "target_hours": target_hours,
        "target_intensity": target_intensity,
        "zero_examples": [
            {
                "model": item["model"],
                "row": item["row"],
                "route_start": item["route_start"],
                "level": item["levels"][0],
                "condition": item["condition_id"],
                "payments": int(item["payments"]),
                "discount": normalize_discount(item["discount"]),
                "total": round_cop(item["total"]),
            }
            for item in zero_records
        ],
        "sample_keys": [list(key) for key in keys[:3]],
        "excel_sha256": hashlib.sha256(BOOK.read_bytes()).hexdigest().upper(),
    }
    if errors:
        write_error_report(errors)
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        raise SystemExit(2)

    now = datetime.now(TIME_ZONE)
    updated_at = now.isoformat(timespec="seconds")
    normalized = [normalized_record(item, updated_at) for item in records]
    try:
        validate_normalized(normalized)
        result["catalogo_normalizado"] = {
            "total": len(normalized),
            "idiomas": dict(Counter(item["idioma_id"] for item in normalized)),
            "modalidades": dict(Counter(item["modelo_tarifario_id"] for item in normalized)),
        }
        if "--auditar" in sys.argv:
            result["publicacion"] = {"omitida": True, "motivo": "Modo de auditoría de solo lectura."}
        else:
            metadata = {
                "producto": "Smart Instituto — Modelo MP",
                "version": MODEL_VERSION,
                "actualizado_en": updated_at,
                "actualizado_en_mostrado": now.strftime("%d/%m/%Y %I:%M %p"),
                "aplicacion_nacional": True,
                "total_registros": len(normalized),
                "registros_por_idioma": dict(Counter(item["idioma_id"] for item in normalized)),
                "registros_por_modalidad": dict(Counter(item["modelo_tarifario_id"] for item in normalized)),
                "registros_por_condicion": dict(Counter(item["condicion_id"] for item in normalized)),
                "archivo_origen": BOOK.name,
                "excel_sha256": result["excel_sha256"],
                "hojas_autorizadas": list(SHEETS),
                "registros_institucionales_intactos": 750,
                "total_catalogos_disponibles": 1202,
            }
            result["publicacion"] = publish_catalog(normalized, metadata, now)
            if ERROR_REPORT.exists():
                ERROR_REPORT.unlink()
    except Exception as error:
        errors.append(str(error))
        result["status"] = "BLOQUEADO"
        result["errors"] = errors
        write_error_report(errors)
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        raise SystemExit(2) from error

    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
